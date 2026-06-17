"""Tests for unified monthly report (Gate 3 + Gate 2 parallel run)."""

from datetime import date
from pathlib import Path

import pandas as pd
import pytest

from app.services.monthly_unified import analyze_unified, build_unified_export_workbook
from app.services.monthly_consolidated import result_from_metrics_dict


@pytest.fixture
def unified_paths(tmp_path):
    ql = pd.DataFrame(
        {
            "Lead Source": ["Cold Calling", "Direct Mail"],
            "Create Date": ["2025-03-10", "2025-03-15"],
            "Street": ["100 Main St", "500 Maple Dr"],
            "City": ["Springfield", "Springfield"],
            "State/Province": ["NY", "NY"],
            "Zip/Postal Code": ["11772", "11772"],
        }
    )
    reisift = pd.DataFrame(
        {
            "Property address": ["100 Main St", "500 Maple Dr", "200 Oak Ave"],
            "Property city": ["Springfield", "Springfield", "Springfield"],
            "Property state": ["NY", "NY", "NY"],
            "Property zip": ["11772", "11772", "11772"],
            "Created": ["2025-03-01", "2025-03-01", "2025-03-01"],
            "Lists": ["High Equity", "High Equity", "High Equity"],
            "Tags": [
                "List Purchased 8020 1/2025,(8020) CC - 2/2025",
                "List Purchased 8020 2/2025,(8020) SMS - 3/2025",
                "List Purchased 8020 1/2025,(8020) DM - 2/2025",
            ],
        }
    )
    closings = pd.DataFrame(
        {
            "Street": ["200 Oak Ave"],
            "City": ["Springfield"],
            "State": ["NY"],
            "Zip": ["11772"],
            "Date Closed": ["2025-06-01"],
            "Stage": ["Closed"],
        }
    )
    ql_path = tmp_path / "ql.csv"
    reisift_path = tmp_path / "reisift.csv"
    closings_path = tmp_path / "closings.xlsx"
    ql.to_csv(ql_path, index=False)
    reisift.to_csv(reisift_path, index=False)
    closings.to_excel(closings_path, index=False)
    return str(ql_path), str(reisift_path), str(closings_path)


def test_analyze_unified_returns_both_reports(unified_paths):
    ql_path, reisift_path, closings_path = unified_paths
    result = analyze_unified(
        ql_path,
        reisift_path,
        closings_path,
        start_date=date(2025, 3, 1),
        end_date=date(2025, 6, 30),
    )
    api = result.to_api_dict()
    assert api["metrics"]["report_type"] == "marketing_ramp"
    assert "total_touch_counts" in api["metrics"]
    assert len(api["rows"]) >= 2
    assert api["consolidated"]["metrics"]["report_type"] == "monthly_consolidated"
    assert len(api["consolidated"]["metrics"].get("lists", [])) >= 1


def test_unified_xlsx_export_includes_ramp_sheet(unified_paths):
    ql_path, reisift_path, closings_path = unified_paths
    result = analyze_unified(
        ql_path,
        reisift_path,
        closings_path,
        start_date=date(2025, 3, 1),
        end_date=date(2025, 6, 30),
    )
    xlsx = build_unified_export_workbook(result.consolidated, result.marketing_ramp.rows)
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(xlsx))
    assert "Marketing Ramp" in wb.sheetnames
    assert "List Performance" in wb.sheetnames


def test_result_from_metrics_dict_roundtrip(unified_paths):
    ql_path, reisift_path, closings_path = unified_paths
    result = analyze_unified(
        ql_path,
        reisift_path,
        closings_path,
        start_date=date(2025, 3, 1),
        end_date=date(2025, 6, 30),
    )
    metrics = result.consolidated.to_dict()
    rebuilt = result_from_metrics_dict(metrics)
    assert rebuilt.cohort_rows == result.consolidated.cohort_rows
    assert len(rebuilt.lists) == len(result.consolidated.lists)
